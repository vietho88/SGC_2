from django.shortcuts import render, redirect
from django.views import View
from django.http import HttpResponse,HttpResponseRedirect,JsonResponse, FileResponse
from Service.decorators import allowed_user
from django.utils.decorators import method_decorator
from Service.models import *
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from Service.forms import *
from django.urls import reverse
import  json
from django.contrib.auth.hashers import make_password, check_password
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.contrib import messages
from django.views.generic import ListView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db import  connection
from django.contrib.auth.decorators import user_passes_test
import pandas as pd
from io import BytesIO
import os
from openpyxl import load_workbook
from django.conf import  settings
import pandas as pd

class UserView(View):
    @method_decorator(user_passes_test(lambda u: u.is_superuser))
    def get(self, request):
        find_cus = 'Tất cả chi nhánh'
        find_name = request.GET.get('input_search_user')
        find_site = Site.objects.first().id
        find_role = Role.objects.filter(site=find_site).values_list('id', flat=True)
        list_user = UserCoop.objects.all().select_related('cus__site', 'role').prefetch_related('manager_cus').order_by('-id')
        is_search_expend = 'false'
        # if len(request.GET) > 1 or (len(request.GET) ==1 and  request.GET.get('page','') ==''):
        if find_name is not None    :
            is_search_expend = request.GET.get('hidden_is_search_expend')

            ###serch nâng cao
            if is_search_expend  == 'true' :
                find_cus = request.GET.get('select_branch_filter_admin')
                find_role = request.GET.getlist('select_role_filter_admin')
                find_site = request.GET.get('select_site')
                if find_cus == 'all':
                    list_user = UserCoop.objects.filter(role__in=find_role, username__contains=find_name).select_related('cus__site', 'role').prefetch_related('manager_cus').order_by('-id')
                else:
                    list_user = UserCoop.objects.filter(role__in=find_role, cus=find_cus, username__contains=find_name).select_related('cus__site', 'role').prefetch_related('manager_cus').order_by('-id')
                    find_cus = int(find_cus)
            else:
                list_user = UserCoop.objects.filter(username__contains=find_name).select_related('cus__site').prefetch_related('manager_cus', 'role').order_by('-id')

        list_site = Site.objects.all()
        list_role = Role.objects.filter(site=find_site)
        list_cus = ListCus.objects.filter(site=find_site)
        page_list_user = Paginator(list_user, settings.PAGINATOR_NUMBER)
        page_number = request.GET.get('page',1)
        page_return = page_list_user.get_page(page_number)
        context = {
            'list_user' : page_return,
            'cus_chosed' : find_cus,
            'role_chosed' : [int(x) for x in find_role],
            'site_chosed' : int(find_site),
            'list_cus' : list_cus,
            'list_site' : list_site,
            'list_role' : list_role,
            'name_search' : find_name,
            'is_search_expend' : is_search_expend,
        }
        return render(request, 'admin/user.html', context)


class UserAddView(View):
    @method_decorator(user_passes_test(lambda u: u.is_superuser))
    def get(self, request):
        form_add_user = UserCreateForm
        form_edit_user = UserEditForm
        list_site = Site.objects.all().values('id', 'name')
        context = {
            'form_add_user': form_add_user,
            'form_edit_user': form_edit_user,
            'list_site': list_site
        }
        return render(request, 'admin/add_user.html', context)

    @method_decorator(user_passes_test(lambda u: u.is_superuser))
    def post(self, request):
        ###Add new user###
        message_error = []
        data_add_user = UserCreateForm(request.POST)

        role = request.POST.get('role','')
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        email = request.POST.get('email', None)
        phone_number = request.POST.get('phone_number', None)
        cus = request.POST.get('cus', '')
        site = request.POST.get('site', '')
        # type_ttpp = request.POST.get('type_ttpp', '')
        # is_ttpp = request.POST.get('ttpp', '')
        is_take_photo = request.POST.get('is_take_photo', False)
        cus_manager = request.POST.getlist('cus_manager', [])

        if username == '':
            message_error.append('Tên người dùng không được phép trống')
            messages.error(request, 'Tên người dùng không được phép trống')
        else:
            regex_username = '^[a-z,0-9,_,-]{6,}$'
            if not re.search(regex_username, username):
                message_error.append('Tên người dùng phải gồm chữ viết thường hoặc số (có thể chứa _, -) , có độ dài lớn hơn 6 ký tự')
                messages.error(request, 'Tên người dùng phải gồm chữ viết thường hoặc số (có thể chứa _, -) , có độ dài lớn hơn 6 ký tự')
        if UserCoop.objects.filter(username=username).exists():
            message_error.append('Tên người dùng này đã tồn tại trên hệ thống')
            messages.error(request, 'Tên người dùng này đã tồn tại trên hệ thống')
        if email :
            regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
            if not re.search(regex, email):
                message_error.append('Email Không đúng định dạng')
                messages.error(request, 'Email Không đúng định dạng')
        if phone_number:
            regex_phone = '^[0-9]{10,13}$'
            if not re.search(regex_phone, phone_number):
                message_error.append('Số điện thoại không đúng định dạng(10 - 13 số)')
                messages.error(request, 'Số điện thoại không đúng định dạng(10 - 13 số)')
        if message_error:
            form_add_user = data_add_user
            list_site = Site.objects.all().values('id', 'name')
            find_role = Role.objects.filter(site_id=site)
            find_cus = ListCus.objects.filter(site_id=site)
            find_cus_manager = ListCus.objects.all().select_related('site').order_by('site')
            context = {
                'error' : 'has_error',
                'site_id': site,
                'cus_id': cus,
                'role_id': role,
                'form_add_user': form_add_user,
                'cus_manager': cus_manager,
                'find_role': find_role,
                'find_cus': find_cus,
                'find_cus_manager': find_cus_manager,
                'list_site': list_site
            }
            return render(request, 'admin/add_user.html', context)

        new_user = UserCoop.objects.create(
            username=username, password=make_password(username), email=email,
            phone_number=phone_number, cus_id=cus, role_id=role, is_take_photo = is_take_photo)
        manager_cus = ListCus.objects.filter(id__in=cus_manager)
        new_user.manager_cus.add(*manager_cus)
        messages.success(request, 'success')
        return redirect(reverse('admin_user'))


class UserEditView(View):
    @method_decorator(user_passes_test(lambda u: u.is_superuser))
    def get(self, request,id):
        user = UserCoop.objects.filter(pk=id).first()
        site_id = user.cus.site.id
        form_edit_user = UserEditForm(initial={
            'username' : user.username,
            'email' : user.email,
            'phone_number' : user.phone_number,
            'type_ttpp' : user.type_ttpp,
            'site' : site_id,
            'is_take_photo' : user.is_take_photo
        })
        find_role = Role.objects.filter(site= site_id)
        find_cus = ListCus.objects.filter(site= site_id)
        find_cus_manager = ListCus.objects.all().select_related('site').order_by('site')
        cus_manager =  [i.id for i in user.manager_cus.all()]
        context = {
            'cus_id' : user.cus.id,
            'role_id' : user.role.id,
            'form_edit_user': form_edit_user,
            'cus_manager': cus_manager,
            'find_role': find_role,
            'find_cus': find_cus,
            'find_cus_manager': find_cus_manager,
            'id' : id
        }
        return render(request, 'admin/edit_user.html',context)
    def post(self, request, id):
        message_error = []
        data_add_user = UserEditForm(request.POST)

        role = request.POST.get('role', '')
        username = request.POST.get('username', '')
        email = request.POST.get('email', None)
        phone_number = request.POST.get('phone_number', None)
        cus = request.POST.get('cus', '')
        is_take_photo = request.POST.get('is_take_photo', False)
        cus_manager = request.POST.getlist('cus_manager', [])

        if username == '':
            message_error.append('Tên người dùng không được phép trống')
            messages.error(request, 'Tên người dùng không được phép trống')
        else:
            regex_username = '^[a-z0-9,_,-]{6,}$'
            if not re.search(regex_username, username):
                message_error.append('Tên người dùng phải gồm chữ viết thường hoặc số (có thể chứa _, -) , có độ dài lớn hơn 6 ký tự')
                messages.error(request, 'Tên người dùng phải gồm chữ viết thường hoặc số (có thể chứa _, -) , có độ dài lớn hơn 6 ký tự')
        if UserCoop.objects.filter(username=username).exclude(pk=id).exists():
            message_error.append('Tên người dùng này đã tồn tại trên hệ thống')
            messages.error(request, 'Tên người dùng này đã tồn tại trên hệ thống')
        if email :
            regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
            if not re.search(regex, email):
                message_error.append('Email Không đúng định dạng')
                messages.error(request, 'Email Không đúng định dạng')
        if phone_number :
            regex_phone = '^[0-9]{10,13}$'
            if not re.search(regex_phone, phone_number):
                message_error.append('Số điện thoại không đúng định dạng, Số điện thoại phải gồm 10 - 13 số')
                messages.error(request, 'Số điện thoại không đúng định dạng, Số điện thoại phải gồm 10 - 13 số')

        if message_error:
            return HttpResponseRedirect(request.META.get("HTTP_REFERER"))

        UserCoop.objects.filter(pk=id).update(
            username=username, email=email,
            phone_number=phone_number, cus_id=cus, role_id=role, is_take_photo = is_take_photo)
        manager_cus = ListCus.objects.filter(id__in=cus_manager)
        edit_user = UserCoop.objects.get(pk=id)
        edit_user.manager_cus.clear()
        edit_user.manager_cus.add(*manager_cus)
        messages.success(request, 'Sửa thông tin người dùng thành công')
        return HttpResponseRedirect(request.META.get("HTTP_REFERER"))

# @csrf_exempt
@login_required(login_url='/login')
@user_passes_test(lambda u: u.is_superuser)
def reset_pass(request):
    if request.method == 'POST':
        message = request.POST.get('message', '')
        id = request.POST.get('id')
        message_error = ''
        message_success = ''
        lock_status = None
        user = UserCoop.objects.filter(id=id)
        if user.exists():
            if message == 'reset':
                UserCoop.objects.filter(id=id).update(password=make_password('Saigoncoop'))
                message_success = 'Success'
            elif message == 'lock':
                old_lock_status = user[0].is_lock
                if old_lock_status :
                    user.update(is_lock = False)
                    lock_status = 0
                else:
                    user.update(is_lock = True)
                    lock_status = 1
                message_success ='Success'
        else:
            message_error = 'Error. Not find id!'

    return JsonResponse({
        'message_error' : message_error,
        'message_success' : message_success,
        'lock_status' : lock_status
    }, safe=False)


@user_passes_test(lambda u: u.is_superuser)
def export_excel_user(request):
    data_export = []
    find_name = request.GET.get('input_search_user')
    if find_name is not None:
        ###serch nâng cao
        find_cus = request.GET.get('select_branch_filter_admin')
        find_role = request.GET.getlist('select_role_filter_admin')
        if find_cus == 'all':
            list_user = UserCoop.objects.filter(
                role__in=find_role,
                username__contains=find_name
                ).select_related('cus__site', 'role').prefetch_related('manager_cus')
        else:
            list_user = UserCoop.objects.filter(
                role__in=find_role,
                cus=find_cus,
                username__contains=find_name
                ).select_related('cus__site', 'role').prefetch_related('manager_cus')
            find_cus = int(find_cus)
    else:
        list_user = UserCoop.objects.all().select_related('cus__site', 'role').prefetch_related('manager_cus')

    for index, user in enumerate(list_user,start=1):
        dict_user =  [
            index,
            user.username,
            'Khóa' if user.is_lock else 'Không Khóa',
            user.role.name,
            user.email,
            user.cus.name,
            user.phone_number,
            user.cus.address,
            list(user.manager_cus.all().values_list('name', flat=True))
        ]
        data_export.append(dict_user)

    column = ['STT','Tên người dùng', 'Trạng thái tài khoản', 'Vai trò', 'Email', 'Đơn vị' , 'Số điện thoại', 'Địa Chỉ', 'Quản lý các đơn vị']
    df = pd.DataFrame(data=data_export, columns=column)
    with BytesIO() as b:
    # Use the StringIO object as the filehandle.
        writer = pd.ExcelWriter( b, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()
        response = HttpResponse(b.getvalue(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename= xuat_danh_sach_nguoi_dung.xlsx'
        return response

@user_passes_test(lambda u : u.is_superuser )
def import_excel_user(request):
    if request.method == 'POST':
        dict_role = {}
        dict_cus = {}
        list_fail = []
        list_success = []

        file = request.FILES.get('file')
        site = request.POST.get('site', 1)
        pd_file = pd.read_excel(file, header=0, usecols='A:H', engine='openpyxl')
        pd_file = pd_file.head(1000)
        pd_file = pd_file.dropna()
        pd_list_add = pd_file.values
        roles = Role.objects.filter(site_id = site)

        for x in roles:
            dict_role[x.name] = x.id
        cuses = ListCus.objects.all()
        for x in cuses:
            dict_cus[x.name] = x.id
        for row in pd_list_add:
            role = dict_role[row[2]]
            user_name = row[1].strip()
            regex_username = '^[a-z,0-9,_,-]{6,}$'
            if not re.search(regex_username, user_name) :
                list_fail.append([user_name, 'Tên người dùng không đúng định dạng'])
                continue
            if UserCoop.objects.filter(username=user_name).exists():
                list_fail.append([user_name, 'Tên người dùng đã tồn tại trên hệ thống'])
                continue
            email = row[3]
            regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
            if not re.search(regex, email) :
                list_fail.append([user_name, 'Email Không đúng định dạng'])
                continue
            cus = dict_cus[row[4]]
            try:
                phone_number =  "0" + str(row[5]).split('.')[0]
            except:
                phone_number = row[5]
            regex_phone = '^[0-9]{10,13}$'
            if not re.search(regex_phone, str(phone_number)):
                list_fail.append([user_name, 'Số điện thoại không đúng định dạng, Số điện thoại phải gồm 10 - 13 số'])
                continue
            cus_manager = row[6].split(', ')
            is_take_photo = 1 if row[7] == 'có' else 0
            new_user = UserCoop.objects.create(
                username=user_name, password=make_password(user_name), email=email,
                phone_number=phone_number, cus_id=cus, role_id=role, is_take_photo=is_take_photo)
            manager_cus = ListCus.objects.filter(name__in=cus_manager)
            new_user.manager_cus.add(*manager_cus)
            list_success.append(user_name)
        request.session['list_import_user_excel_fail'] =  list_fail
        request.session['list_import_user_excel_success'] =  list_success
        return JsonResponse({
            'message' : 'oke',
            'count_success' : len(list_success),
            'count_fail' : len(list_fail)
        }, safe=False)

@user_passes_test(lambda u : u.is_superuser )
def import_excel_user_dowload_file(request):
    type = request.GET.get('type')
    if type == 'dowload_report':
        list_success = request.session['list_import_user_excel_success']
        list_fail = request.session['list_import_user_excel_fail']
        column = ['Tên người dùng', 'Lỗi']
        df = pd.DataFrame(data=list_fail, columns=column)
        df2 = pd.DataFrame(data=list_success, columns=['Tên người dùng thêm thành công'])
        with BytesIO() as b:
            # Use the StringIO object as the filehandle.
            writer = pd.ExcelWriter(b, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Danh sách thêm thất bại', index=False)
            df2.to_excel(writer, sheet_name='Danh sách thêm thành công', index=False)
            writer.save()
            response = HttpResponse(b.getvalue(),  content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename= bao_cao_tai_len_nguoi_dung.xlsx'
            return response

    site = request.GET.get('site', 1)
    try:
        path_template = settings.MEDIA_ROOT + "/Mau_Import_Nguoi_Dung_Coopapp_2.0.xlsm"
        path_template_return = settings.MEDIA_ROOT + "/Mau_Import_Nguoi_Dung_Coopapp_2020.xlsm"
        wb = load_workbook(filename=path_template, read_only=False, keep_vba=True)
        ws = wb.active
        ws = wb.worksheets[0]
        list_cus_name = []

        list_cus_all = ListCus.objects.all().values('id', 'name')
        list_cus_site = ListCus.objects.filter(site_id = site).values('id', 'name')
        list_role = Role.objects.filter(site_id = site).values('id', 'name')
        row = 1006
        col = 19
        for item in list_cus_all:
            ws.cell(row, col, value=item['name'])
            # ws.cell(row, col - 1, value=item['name'])
            row += 1
        row = 1006
        col_role = 15
        for item in list_role:
            ws.cell(row, col_role, value=item['name'])
            row += 1

        row = 1006
        col_role = 18
        for item in list_cus_site:
            ws.cell(row, col_role, value=item['name'])
            row += 1
        wb.save(path_template_return)
        with open(path_template_return, "rb") as excel:
            data = excel.read()
        response = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename= mau_import_user_excel.xlsm'
        return response
    except Exception as e:
        print(e)
        return HttpResponse(json.dumps({"message": "fail"}), content_type="application/json")
